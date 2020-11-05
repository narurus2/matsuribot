import asyncio,discord,os
from openpyxl import load_workbook
from discord.ext import commands

#토큰
token_path = os.path.dirname( os.path.abspath( __file__ ))+"/token.txt" #텍스트 파일로 부터 토큰 불러옴
t = open(token_path,"r",encoding="utf-8")
token = t.read().split()[0] #토큰을 변수에 저장
#print("Token_key : ",token)

#봇의 설정
game = discord.Game("!도움") #코멘트 설정
bot = commands.Bot(command_prefix='!',status=discord.Status.online,activity=game) #봇의 상태 설정

#봇 시작
@bot.event
async def on_ready(): #봇이 가동시작 되었을때 
    print("마츠리봇 가동 시작") 

@bot.command() #명령어를 추가할때마다 추가해야해
async def 도움(ctx): 
    await ctx.send("무엇을 도와줄까?")
''' 
@bot.command() 
async def 발없찐(ctx):
    await ctx.send("쿠루미봇")
'''
@bot.command() #embed를 이용해서 채팅에 다채로움을 추가함
async def 인사말(ctx):
    embed=discord.Embed(title= f"여기는 쿠루미봇 디스코드입니다.", description=f"다른말로는 노루수용소라고 하죠.", color=0xf3bb76)
    embed.add_field(name=f"우리 서버의 목적은?", value=f"클랜전 비성실 참여자를 기록하기 위해서 만들어졌습니다.", inline=False)
    await ctx.send(embed=embed)

#data_only=Ture로 해줘야 수식이 아닌 값으로 받아온다.
load_wb = load_workbook("/Users/dldud/Desktop/matsuribot/test.xlsx", data_only=True)
#시트 이름으로 불러오기
load_ws = load_wb['9월클랜전']

test_excel = [ "등수 :", load_ws['A8'].value, "클랜명 :", load_ws['B8'].value, "총점수 :", load_ws['C8'].value ]
#test_excel.remove("'")
#test_excel.remove('.')

'''
@bot.command() #명령어를 추가할때마다 추가해야해
async def 테스트(ctx): 
    for x in test_excel:
        await ctx.send(x)
'''
#여기부터 코드가 좀 난잡함
i = 0
test_excel1 = []
test_excel2 = []
for x in (range(len(test_excel))):
    if i%2 == 0:
        test_excel1.append(test_excel[x])
    else:
        test_excel2.append(test_excel[x])
    i += 1
#print(test_excel1) 배열 분배 확인용
#print(test_excel2)

@bot.command() #embed를 이용해서 채팅에 다채로움을 추가함
async def 테스트2(ctx):
    i = 0
    embed=discord.Embed(title=f"테스트1", description=f"테스트2", color=0xf3bb76)
    for x in test_excel1:
        y = test_excel2[i]
        embed.add_field(name=x, value=y, inline=False)
        i += 1
    await ctx.send(embed=embed)

#셀 주소로 값 출력
#print("등수 :", load_ws['A8'].value, "클랜명 :", load_ws['B8'].value,"총점수 :", load_ws['C8'].value)

#셀 좌표로 값 출력
#print(load_ws.cell(1,8).value)

access_token = os.environ["BOT_TOKEN"]
bot.run(token) 











